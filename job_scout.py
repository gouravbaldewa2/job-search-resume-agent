#!/usr/bin/env python3
"""Job scouting agent: search LinkedIn, score jobs, generate Excel tracker."""

import json
import os
import re
import time
import subprocess
from datetime import datetime, timedelta
from bs4 import BeautifulSoup

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

TODAY = datetime.now().strftime("%Y-%m-%d")
OUTPUT_DIR = f"output/{TODAY}"
SEEN_JOBS_FILE = "seen_jobs.json"

TITLES = [
    "lead%20product%20manager",
    "staff%20product%20manager",
    "principal%20product%20manager",
    "group%20product%20manager",
    "senior%20product%20manager",
]

TITLE_FILTERS = [
    "lead product manager",
    "staff product manager",
    "principal product manager",
    "group product manager",
    "senior product manager",
    "sr. product manager",
    "sr product manager",
]

LOCATIONS = [
    {"name": "Bangalore", "query": "Bangalore%2C%20India", "remote_only": False},
    {"name": "Dubai", "query": "Dubai%2C%20United%20Arab%20Emirates", "remote_only": False},
    {"name": "Netherlands", "query": "Netherlands", "remote_only": False},
    {"name": "Remote India", "query": "India", "remote_only": True},
]

MAX_JOBS_PER_LOCATION = 10
MIN_REMOTE_INDIA = 10


def load_seen_jobs():
    """Load and prune seen_jobs.json."""
    if not os.path.exists(SEEN_JOBS_FILE):
        return {}
    with open(SEEN_JOBS_FILE, "r") as f:
        data = json.load(f)
    cutoff = (datetime.now() - timedelta(days=14)).strftime("%Y-%m-%d")
    pruned = {k: v for k, v in data.items() if v.get("date", "") >= cutoff}
    print(f"Loaded {len(data)} seen jobs, pruned to {len(pruned)} (within 14 days)")
    return pruned


def curl_fetch(url, retries=2):
    """Fetch URL content via curl subprocess."""
    for attempt in range(retries + 1):
        try:
            result = subprocess.run(
                ["curl", "-s", "-L",
                 "-H", "User-Agent: Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                 "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                 "-H", "Accept-Language: en-US,en;q=0.5",
                 url],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout
        except subprocess.TimeoutExpired:
            pass
        if attempt < retries:
            time.sleep(2 * (attempt + 1))
    return ""


def extract_job_id(element):
    """Extract job ID from HTML element."""
    # Try data-entity-urn first
    urn = element.get("data-entity-urn", "")
    if urn:
        match = re.search(r"jobPosting:(\d+)", urn)
        if match:
            return match.group(1)
    # Check inner divs
    for div in element.find_all("div", attrs={"data-entity-urn": True}):
        urn = div.get("data-entity-urn", "")
        match = re.search(r"jobPosting:(\d+)", urn)
        if match:
            return match.group(1)
    # Fallback: extract from href
    link = element.find("a", href=True)
    if link:
        href = link["href"]
        match = re.search(r"(\d{7,})", href)
        if match:
            return match.group(1)
    return None


def parse_job_listings(html):
    """Parse LinkedIn job listing HTML into structured data."""
    jobs = []
    if not html or len(html.strip()) < 50:
        return jobs
    soup = BeautifulSoup(html, "html.parser")

    # Try multiple selectors for job cards
    cards = soup.find_all("li")
    if not cards:
        cards = soup.find_all("div", class_=re.compile(r"job|result|card", re.I))

    for card in cards:
        # Extract title
        title_el = card.find(["h3", "h4", "a"], class_=re.compile(r"title|name", re.I))
        if not title_el:
            title_el = card.find("a", href=re.compile(r"/jobs/"))
        if not title_el:
            continue
        title = title_el.get_text(strip=True)

        # Extract company
        company_el = card.find(["h4", "span", "a"], class_=re.compile(r"company|subtitle|name", re.I))
        company = company_el.get_text(strip=True) if company_el else "Unknown"

        # Extract location
        location_el = card.find(["span", "div"], class_=re.compile(r"location|locale", re.I))
        city = location_el.get_text(strip=True) if location_el else ""

        # Extract URL
        link_el = card.find("a", href=True)
        url = link_el["href"] if link_el else ""
        if url and not url.startswith("http"):
            url = "https://www.linkedin.com" + url

        # Extract job ID
        job_id = extract_job_id(card)
        if not job_id and url:
            match = re.search(r"(\d{7,})", url)
            if match:
                job_id = match.group(1)

        if title and job_id:
            jobs.append({
                "title": title,
                "company": company,
                "city": city,
                "url": url,
                "job_id": job_id,
            })
    return jobs


def title_matches(title):
    """Check if job title matches our target roles."""
    t = title.lower().strip()
    return any(f in t for f in TITLE_FILTERS)


def search_jobs_for_location(location, seen_jobs):
    """Search all titles for a given location."""
    all_jobs = []
    skipped = []
    seen_ids = set()

    is_remote = location["remote_only"]
    loc_name = location["name"]
    loc_query = location["query"]

    for title_query in TITLES:
        url = (
            f"https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search"
            f"?keywords={title_query}&location={loc_query}&f_TPR=r86400"
        )
        if is_remote:
            url += "&f_WT=2"

        # Fetch multiple pages for Remote India to hit minimum 10
        pages_to_try = 3 if (is_remote and loc_name == "Remote India") else 2
        for page in range(pages_to_try):
            page_url = url + f"&start={page * 25}"
            print(f"  Fetching: {title_query.replace('%20', ' ')} in {loc_name} (page {page})...")
            html = curl_fetch(page_url)
            if not html:
                print(f"    No response / blocked")
                continue

            jobs = parse_job_listings(html)
            print(f"    Found {len(jobs)} raw listings")

            for job in jobs:
                jid = job["job_id"]
                if jid in seen_ids:
                    continue
                seen_ids.add(jid)

                if not title_matches(job["title"]):
                    continue

                if jid in seen_jobs:
                    skipped.append({
                        "title": job["title"],
                        "company": job["company"],
                        "date": seen_jobs[jid].get("date", "unknown"),
                    })
                    print(f"    Skipped (seen on {seen_jobs[jid].get('date', '?')}): {job['title']} at {job['company']}")
                    continue

                job["location_group"] = loc_name
                all_jobs.append(job)

            time.sleep(1.5)  # Rate limiting

        if len(all_jobs) >= MAX_JOBS_PER_LOCATION and not (is_remote and len(all_jobs) < MIN_REMOTE_INDIA):
            break

    # Cap results
    cap = max(MAX_JOBS_PER_LOCATION, MIN_REMOTE_INDIA) if is_remote else MAX_JOBS_PER_LOCATION
    return all_jobs[:cap], skipped


def fetch_job_description(job_id):
    """Fetch the full job description for a job ID."""
    url = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{job_id}"
    html = curl_fetch(url)
    if not html:
        return "", []

    soup = BeautifulSoup(html, "html.parser")

    # Extract description text
    desc_el = soup.find("div", class_=re.compile(r"description|content", re.I))
    if not desc_el:
        desc_el = soup.find("section", class_=re.compile(r"description", re.I))
    if not desc_el:
        desc_el = soup  # fallback to entire page

    text = desc_el.get_text(separator="\n", strip=True)

    # Extract keywords
    keywords = extract_keywords(text)
    return text, keywords


def extract_keywords(text):
    """Extract key domain/skill keywords from JD text."""
    keyword_patterns = [
        r"A/B test", r"experimentation", r"growth", r"conversion",
        r"marketplace", r"consumer", r"fintech", r"e-commerce", r"ecommerce",
        r"platform", r"mobile", r"agile", r"scrum", r"stakeholder",
        r"cross-functional", r"data[- ]driven", r"analytics", r"SQL",
        r"machine learning", r"ML", r"AI", r"artificial intelligence",
        r"product strategy", r"roadmap", r"product lifecycle",
        r"user experience", r"UX", r"retention", r"engagement",
        r"payments?", r"lending", r"banking", r"insurance",
        r"logistics", r"mobility", r"transportation",
        r"B2B", r"B2C", r"SaaS", r"API",
        r"segmentation", r"personali[sz]ation", r"funnel",
        r"KPI", r"OKR", r"metrics", r"revenue",
        r"regulated", r"compliance", r"international",
        r"multi[- ]market", r"global",
    ]
    found = []
    text_lower = text.lower()
    for pattern in keyword_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            # Clean pattern for display
            clean = pattern.replace(r"[- ]", "-").replace(r"?", "").replace(r"[sz]", "s")
            if clean not in found:
                found.append(clean)
    return found[:8]


def score_job(jd_text, keywords):
    """Score a job 1-10 based on relevance to Gourav's profile."""
    text_lower = jd_text.lower()
    score = 0

    # Consumer-facing, marketplaces, two-sided platforms: +3
    if any(w in text_lower for w in ["consumer-facing", "consumer facing", "marketplace",
                                      "two-sided", "2-sided", "consumer product"]):
        score += 3

    # Growth, experimentation, A/B testing, conversion optimisation: +2
    if any(w in text_lower for w in ["a/b test", "experimentation", "growth product",
                                      "conversion optim", "growth strategy", "funnel"]):
        score += 2

    # Fintech with consumer lens: +2
    if any(w in text_lower for w in ["fintech", "digital banking", "lending", "deposits",
                                      "neobank", "consumer finance"]):
        if not any(w in text_lower for w in ["payment rail", "payment infrastructure",
                                              "payment gateway", "payment processing"]):
            score += 2

    # E-commerce, retail, catalog/discovery: +2
    if any(w in text_lower for w in ["e-commerce", "ecommerce", "retail", "catalog",
                                      "catalogue", "product discovery", "shopping"]):
        score += 2

    # Mobility, transportation, logistics with consumer angle: +1
    if any(w in text_lower for w in ["mobility", "transportation", "logistics", "ride",
                                      "delivery", "fleet"]):
        score += 1

    # Platform products serving product teams: +1
    if any(w in text_lower for w in ["platform product", "internal platform", "developer platform",
                                      "product platform"]):
        score += 1

    # AI/ML product management: +1
    if any(w in text_lower for w in ["ai product", "ml product", "machine learning",
                                      "artificial intelligence", "ai/ml"]):
        score += 1

    # International/multi-market: +1
    if any(w in text_lower for w in ["international", "multi-market", "global market",
                                      "multi-country", "cross-border"]):
        score += 1

    # Stakeholder management, cross-functional leadership: +1
    if any(w in text_lower for w in ["stakeholder", "cross-functional", "cross functional"]):
        score += 1

    # Regulated environments: +1
    if any(w in text_lower for w in ["regulated", "compliance", "regulatory"]):
        score += 1

    # Subtract: heavy engineering focus: -2
    if any(w in text_lower for w in ["system design", "system architecture",
                                      "distributed systems", "infrastructure engineer"]):
        if not any(w in text_lower for w in ["product strategy", "product vision", "user"]):
            score -= 2

    # Subtract: niche domain: -1
    if any(w in text_lower for w in ["phd required", "md required",
                                      "years in healthcare", "clinical experience"]):
        score -= 1

    return max(1, min(10, score))


def should_exclude(jd_text, company):
    """Check if job should be excluded. Returns (bool, reason)."""
    text_lower = jd_text.lower()
    company_lower = company.lower()

    # Payment processing / gateway / infrastructure
    payment_terms = ["payment processing", "payment gateway", "payment infrastructure",
                     "payment rails", "payment orchestration"]
    if any(t in text_lower for t in payment_terms):
        # Check if it's primarily about payments
        payment_count = sum(1 for t in payment_terms if t in text_lower)
        if payment_count >= 2 or (payment_count >= 1 and "payment" in (jd_text.lower().split("responsibilities")[0] if "responsibilities" in text_lower else text_lower[:500])):
            return True, "Payment processing/gateway/infrastructure focus"

    # Pure B2B SaaS
    b2b_signals = ["enterprise software", "b2b saas", "enterprise saas",
                   "sell to businesses", "enterprise customers only"]
    consumer_signals = ["consumer", "b2c", "end user", "customer-facing", "consumer-facing"]
    b2b_count = sum(1 for s in b2b_signals if s in text_lower)
    consumer_count = sum(1 for s in consumer_signals if s in text_lower)
    if b2b_count >= 2 and consumer_count == 0:
        return True, "Pure B2B SaaS with no consumer-facing component"

    # Staffing/recruitment agency
    staffing_terms = ["staffing", "recruitment agency", "consulting firm",
                      "on behalf of our client", "hiring for our client"]
    if any(t in text_lower for t in staffing_terms) or any(
        t in company_lower for t in ["staffing", "recruiting", "manpower", "randstad",
                                      "adecco", "hays", "michael page", "robert half"]):
        return True, "Staffing/recruitment agency"

    # Deep niche domain
    niche_terms = ["5+ years in healthcare compliance", "phd in ml required",
                   "medical device experience required", "pharma experience required",
                   "cpa required", "bar admission required"]
    for t in niche_terms:
        if t in text_lower:
            return True, f"Requires niche domain expertise: {t}"

    return False, ""


def create_excel(jobs_data, output_path):
    """Create styled Excel tracker."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    headers = [
        "S.No", "Organization", "Job Title", "Location Group", "City",
        "Relevance Score", "Key JD Keywords", "LinkedIn Job Link",
        "Product Leader (Name)", "Product Leader (Email)"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sort by relevance score descending
    jobs_data.sort(key=lambda x: x.get("score", 0), reverse=True)

    # Write data
    for idx, job in enumerate(jobs_data, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=job.get("company", "")).border = thin_border
        ws.cell(row=row, column=3, value=job.get("title", "")).border = thin_border
        ws.cell(row=row, column=4, value=job.get("location_group", "")).border = thin_border
        ws.cell(row=row, column=5, value=job.get("city", "")).border = thin_border
        ws.cell(row=row, column=6, value=job.get("score", 0)).border = thin_border

        # Keywords
        kw_cell = ws.cell(row=row, column=7, value=", ".join(job.get("keywords", [])))
        kw_cell.border = thin_border
        kw_cell.alignment = Alignment(wrap_text=True)

        # Hyperlink
        url = job.get("url", "")
        if not url and job.get("job_id"):
            url = f"https://www.linkedin.com/jobs/view/{job['job_id']}"
        link_cell = ws.cell(row=row, column=8, value=url)
        link_cell.border = thin_border
        if url:
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row, column=9, value=job.get("leader_name", "N/A")).border = thin_border
        ws.cell(row=row, column=10, value=job.get("leader_email", "N/A")).border = thin_border

    # Column widths
    col_widths = [6, 25, 30, 15, 20, 14, 40, 45, 25, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs_data) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Conditional formatting on Relevance Score (column F)
    score_col = "F"
    data_range = f"{score_col}2:{score_col}{len(jobs_data) + 1}"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add(data_range, CellIsRule(
        operator="greaterThanOrEqual", formula=["8"], fill=green_fill
    ))
    ws.conditional_formatting.add(data_range, CellIsRule(
        operator="between", formula=["5", "7"], fill=yellow_fill
    ))
    ws.conditional_formatting.add(data_range, CellIsRule(
        operator="lessThanOrEqual", formula=["4"], fill=red_fill
    ))

    wb.save(output_path)
    print(f"Excel saved to {output_path}")


def write_summary(all_jobs, filtered_jobs, skipped_jobs, blocked_locations):
    """Write summary.md."""
    lines = [
        f"# Job Scout Summary - {TODAY}\n",
        "## Next step: Generate resumes for selected jobs",
        "Review the Excel tracker and note the S.No values for jobs you want to apply to. Then run:",
        "```",
        "./generate_selected.sh 1 3 5",
        "```\n",
    ]

    # Group jobs by location
    location_groups = {}
    for job in all_jobs:
        loc = job.get("location_group", "Unknown")
        if loc not in location_groups:
            location_groups[loc] = []
        location_groups[loc].append(job)

    lines.append("## New Jobs Found\n")
    if not all_jobs:
        lines.append("No new jobs were found in this run.\n")

    for loc in ["Bangalore", "Dubai", "Netherlands", "Remote India"]:
        jobs = location_groups.get(loc, [])
        lines.append(f"### {loc} ({len(jobs)} jobs)\n")
        if not jobs:
            if loc in blocked_locations:
                lines.append("LinkedIn returned empty results (possibly rate-limited).\n")
            else:
                lines.append("No matching jobs found.\n")
        else:
            for job in sorted(jobs, key=lambda x: x.get("score", 0), reverse=True):
                lines.append(
                    f"- **{job['title']}** at {job['company']} "
                    f"(Score: {job.get('score', 'N/A')}) - {job.get('city', '')}"
                )
            lines.append("")

    # Filtered out
    lines.append("## Filtered Out\n")
    if filtered_jobs:
        for fj in filtered_jobs:
            lines.append(f"- **{fj['title']}** at {fj['company']} - Reason: {fj['reason']}")
    else:
        lines.append("No jobs were filtered out.\n")
    lines.append("")

    # Skipped (already seen)
    lines.append("## Skipped (already seen)\n")
    if skipped_jobs:
        for sj in skipped_jobs:
            lines.append(
                f"- **{sj['title']}** at {sj['company']} (seen on {sj.get('date', 'unknown')})"
            )
    else:
        lines.append("No previously seen jobs were encountered.\n")
    lines.append("")

    # Blocked locations note
    if blocked_locations:
        lines.append("## Notes\n")
        lines.append(
            f"LinkedIn may have rate-limited or blocked requests for: {', '.join(blocked_locations)}\n"
        )

    summary_path = f"{OUTPUT_DIR}/summary.md"
    with open(summary_path, "w") as f:
        f.write("\n".join(lines))
    print(f"Summary written to {summary_path}")
    return summary_path


def main():
    print(f"=== Job Scout - {TODAY} ===\n")

    # Load seen jobs
    seen_jobs = load_seen_jobs()

    # Search all locations
    all_new_jobs = []
    all_skipped = []
    blocked_locations = []

    for location in LOCATIONS:
        print(f"\n--- Searching: {location['name']} ---")
        jobs, skipped = search_jobs_for_location(location, seen_jobs)
        if not jobs and not skipped:
            blocked_locations.append(location["name"])
        all_new_jobs.extend(jobs)
        all_skipped.extend(skipped)
        print(f"  Found {len(jobs)} new jobs, skipped {len(skipped)} seen jobs")

    print(f"\n=== Total new jobs found: {len(all_new_jobs)} ===\n")

    # Deduplicate across locations (same job_id might appear in multiple searches)
    seen_ids = set()
    unique_jobs = []
    for job in all_new_jobs:
        if job["job_id"] not in seen_ids:
            seen_ids.add(job["job_id"])
            unique_jobs.append(job)
    all_new_jobs = unique_jobs
    print(f"After cross-location dedup: {len(all_new_jobs)} unique jobs\n")

    # Fetch JDs and score
    scored_jobs = []
    filtered_jobs = []

    for i, job in enumerate(all_new_jobs):
        print(f"Fetching JD {i+1}/{len(all_new_jobs)}: {job['title']} at {job['company']}...")
        jd_text, keywords = fetch_job_description(job["job_id"])
        time.sleep(1)

        if not jd_text or len(jd_text) < 50:
            print(f"  Could not fetch JD, using title-only scoring")
            jd_text = job["title"] + " " + job.get("company", "")
            keywords = extract_keywords(jd_text)

        # Check exclusions
        excluded, reason = should_exclude(jd_text, job["company"])
        if excluded:
            print(f"  EXCLUDED: {reason}")
            filtered_jobs.append({**job, "reason": reason})
            continue

        # Score
        score = score_job(jd_text, keywords)
        job["score"] = score
        job["keywords"] = keywords
        job["jd_text"] = jd_text
        job["leader_name"] = "N/A"
        job["leader_email"] = "N/A"
        scored_jobs.append(job)
        print(f"  Score: {score}, Keywords: {', '.join(keywords[:5])}")

    print(f"\n=== Scored jobs: {len(scored_jobs)}, Filtered out: {len(filtered_jobs)} ===\n")

    # Generate Excel
    excel_path = f"{OUTPUT_DIR}/job_tracker_{TODAY}.xlsx"
    create_excel(scored_jobs, excel_path)

    # Update seen_jobs
    for job in all_new_jobs:
        seen_jobs[job["job_id"]] = {
            "date": TODAY,
            "company": job.get("company", "Unknown"),
        }
    # Also add filtered jobs
    for job in filtered_jobs:
        if "job_id" in job:
            seen_jobs[job["job_id"]] = {
                "date": TODAY,
                "company": job.get("company", "Unknown"),
            }
    with open(SEEN_JOBS_FILE, "w") as f:
        json.dump(seen_jobs, f, indent=2)
    print(f"Updated {SEEN_JOBS_FILE} with {len(seen_jobs)} entries")

    # Write summary
    write_summary(scored_jobs, filtered_jobs, all_skipped, blocked_locations)

    print("\n=== Done! ===")


if __name__ == "__main__":
    main()
