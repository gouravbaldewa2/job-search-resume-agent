#!/usr/bin/env python3
"""Reorganize output into date-wise folders, cap 10 jobs/location, generate Excel."""

import json
import os
import re
import shutil
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TODAY = "2026-03-02"
OUTPUT_DIR = "output"
DATE_DIR = os.path.join(OUTPUT_DIR, TODAY)
MAX_PER_LOCATION = 10

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def try_find_product_leader(company_name):
    """Try to find head of product / senior product leader from LinkedIn company page.

    LinkedIn's public guest API is very limited for people search.
    We attempt a best-effort search but this will often return empty.
    """
    name = ""
    email = ""

    # Try LinkedIn people search for head of product at the company
    clean_company = re.sub(r"[^a-zA-Z0-9 ]", "", company_name).strip()
    query = f"head%20of%20product%20{clean_company.replace(' ', '%20')}"
    url = (
        f"https://www.linkedin.com/search/results/people/"
        f"?keywords={query}&origin=GLOBAL_SEARCH_HEADER"
    )

    # LinkedIn people search requires authentication, so try the company page instead
    company_slug = re.sub(r"[^a-zA-Z0-9]", "", company_name.lower().replace(" ", ""))
    company_url = f"https://www.linkedin.com/company/{company_slug}"

    try:
        resp = requests.get(company_url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            # Look for any people mentioned on the public company page
            people_links = soup.find_all("a", href=re.compile(r"/in/"))
            for link in people_links[:5]:
                text = link.get_text(strip=True)
                if any(kw in text.lower() for kw in [
                    "product", "chief product", "vp product", "head of product",
                    "director of product", "cpo"
                ]):
                    name = text
                    break
    except Exception:
        pass

    return name, email


def clean_job_link(link):
    """Strip tracking params from LinkedIn job URL for cleaner links."""
    if not link:
        return ""
    # Keep just the base URL up to the job ID
    match = re.match(r"(https?://[^?]+)", link)
    return match.group(1) if match else link


def main():
    # ---- Load all job data ----
    with open("search_result.json") as f:
        search_data = json.load(f)

    with open("/tmp/remote_india_generated.json") as f:
        remote_india_all = json.load(f)

    # Build consolidated job list per location
    all_jobs = {
        "Bangalore": search_data["all_jobs"]["Bangalore"],
        "Dubai": search_data["all_jobs"]["Dubai"],
        "Netherlands": search_data["all_jobs"]["Netherlands"],
        "Remote India": [],
    }

    # For Remote India, use the full list from the supplemental search
    seen_ids = set()
    for job in remote_india_all:
        if job["job_id"] not in seen_ids:
            seen_ids.add(job["job_id"])
            all_jobs["Remote India"].append(job)

    # ---- Cap at MAX_PER_LOCATION ----
    for loc in all_jobs:
        if len(all_jobs[loc]) > MAX_PER_LOCATION:
            print(f"Trimming {loc} from {len(all_jobs[loc])} to {MAX_PER_LOCATION} jobs")
            all_jobs[loc] = all_jobs[loc][:MAX_PER_LOCATION]

    # ---- Reorganize output into date subfolder ----
    print(f"\nReorganizing output/ into output/{TODAY}/")
    os.makedirs(DATE_DIR, exist_ok=True)

    # Build set of filenames we want to keep (based on capped job list)
    keep_filenames = set()
    for loc, jobs in all_jobs.items():
        for job in jobs:
            company = job.get("company", "Unknown")
            company_safe = re.sub(r"[^a-zA-Z0-9]", "", company.replace(" ", ""))[:30]
            if not company_safe:
                company_safe = "Unknown"

            title = job.get("title", "")
            t = title.lower()
            if "lead" in t:
                short = "LeadPM"
            elif "staff" in t:
                short = "StaffPM"
            elif "principal" in t:
                short = "PrincipalPM"
            elif "group" in t:
                short = "GroupPM"
            elif "senior" in t:
                short = "SeniorPM"
            elif "sr" in t:
                short = "PM"
            else:
                short = "PM"

            prefix = "Remote_" if loc == "Remote India" else ""
            fname = f"{prefix}{company_safe}_{short}_{TODAY}.pdf"
            keep_filenames.add(fname)

            # Also check if the filename field is directly available
            if "filename" in job:
                keep_filenames.add(job["filename"])

    # Move kept PDFs to date subfolder
    moved = 0
    for fname in os.listdir(OUTPUT_DIR):
        src = os.path.join(OUTPUT_DIR, fname)
        if not os.path.isfile(src) or not fname.endswith(".pdf"):
            continue
        if fname in keep_filenames:
            dst = os.path.join(DATE_DIR, fname)
            shutil.move(src, dst)
            moved += 1
            print(f"  Moved: {fname}")
        else:
            # Remove PDFs that are over the cap
            os.remove(src)
            print(f"  Removed (over cap): {fname}")

    print(f"Moved {moved} PDFs to {DATE_DIR}")

    # ---- Try to find product leaders ----
    print("\nAttempting to find product leaders (best-effort, limited by public API)...")
    leader_cache = {}
    unique_companies = set()
    for loc, jobs in all_jobs.items():
        for job in jobs:
            unique_companies.add(job.get("company", ""))

    for company in unique_companies:
        if not company:
            continue
        print(f"  Searching for product leader at {company}...")
        name, email = try_find_product_leader(company)
        leader_cache[company] = {"name": name, "email": email}
        if name:
            print(f"    Found: {name}")
        else:
            print(f"    Not found (public data limited)")
        time.sleep(0.3)

    # ---- Generate Excel ----
    print(f"\nGenerating Excel file...")
    wb = Workbook()
    ws = wb.active
    ws.title = f"PM Jobs - {TODAY}"

    # Styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "S.No", "Organization", "Job Title", "Location Group",
        "City", "LinkedIn Job Link", "Resume PDF",
        "Product Leader (Name)", "Product Leader (Email)"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    row_num = 2
    serial = 1
    for loc in ["Bangalore", "Dubai", "Netherlands", "Remote India"]:
        jobs = all_jobs[loc]
        for job in jobs:
            company = job.get("company", "Unknown")
            title = job.get("title", "")
            link = clean_job_link(job.get("link", ""))
            city = job.get("location", "")
            leader_info = leader_cache.get(company, {})
            leader_name = leader_info.get("name", "")
            leader_email = leader_info.get("email", "")

            # Determine PDF filename
            company_safe = re.sub(r"[^a-zA-Z0-9]", "", company.replace(" ", ""))[:30] or "Unknown"
            t = title.lower()
            if "lead" in t:
                short = "LeadPM"
            elif "staff" in t:
                short = "StaffPM"
            elif "principal" in t:
                short = "PrincipalPM"
            elif "group" in t:
                short = "GroupPM"
            elif "senior" in t:
                short = "SeniorPM"
            elif "sr" in t:
                short = "PM"
            else:
                short = "PM"
            prefix = "Remote_" if loc == "Remote India" else ""
            pdf_name = f"{prefix}{company_safe}_{short}_{TODAY}.pdf"

            ws.cell(row=row_num, column=1, value=serial).font = cell_font
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=row_num, column=2, value=company).font = cell_font
            ws.cell(row=row_num, column=3, value=title).font = cell_font
            ws.cell(row=row_num, column=4, value=loc).font = cell_font
            ws.cell(row=row_num, column=5, value=city).font = cell_font

            # LinkedIn link as clickable hyperlink
            link_cell = ws.cell(row=row_num, column=6, value="View Job")
            if link:
                link_cell.hyperlink = link
            link_cell.font = link_font

            ws.cell(row=row_num, column=7, value=pdf_name).font = cell_font
            ws.cell(row=row_num, column=8, value=leader_name if leader_name else "N/A").font = cell_font
            ws.cell(row=row_num, column=9, value=leader_email if leader_email else "N/A").font = cell_font

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).alignment = cell_align
                ws.cell(row=row_num, column=col).border = thin_border

            serial += 1
            row_num += 1

    # Column widths
    col_widths = [6, 30, 45, 15, 30, 15, 40, 25, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num - 1}"

    excel_path = os.path.join(DATE_DIR, f"job_tracker_{TODAY}.xlsx")
    wb.save(excel_path)
    print(f"Excel saved to: {excel_path}")

    # Print summary
    print(f"\n--- Final counts ---")
    total = 0
    for loc, jobs in all_jobs.items():
        print(f"  {loc}: {len(jobs)}")
        total += len(jobs)
    print(f"  Total: {total}")

    return all_jobs


if __name__ == "__main__":
    main()
